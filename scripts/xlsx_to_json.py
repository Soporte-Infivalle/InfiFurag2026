"""
xlsx_to_json.py — Convierte data/preguntas.xlsx a:
  - data/index.json          (lista de módulos y totales, 0.5 KB)
  - data/modulos/<MOD>.json  (preguntas por módulo, avg ~13 KB cada uno)

Ejecutar desde la raíz del proyecto:
    python scripts/xlsx_to_json.py

Requiere: pip install openpyxl
"""

import json, sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: instala openpyxl con:  pip install openpyxl")
    sys.exit(1)

ROOT      = Path(__file__).parent.parent
XLSX_PATH = ROOT / 'data' / 'preguntas.xlsx'
SHEET     = 'Preguntas'

def convert():
    if not XLSX_PATH.exists():
        print(f"ERROR: no se encontró {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: hoja '{SHEET}' no encontrada. Hojas: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET]
    mods = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        opciones = [str(row[i]).strip() for i in range(4, 19)
                    if row[i] and str(row[i]).strip()]
        cod = str(row[1]) if row[1] else ''
        q = {
            'id':      row[0],
            'codigo':  cod,
            'modulo':  cod[:3].upper(),
            'texto':   str(row[2]).strip() if row[2] else '',
            'tipo':    str(row[3]).strip() if row[3] else 'Selección única',
            'opciones': opciones,
        }
        mods[q['modulo']].append(q)

    # Sort each module by ID
    for mod in mods:
        mods[mod].sort(key=lambda q: q['id'])

    # Write per-module JSONs
    mods_dir = ROOT / 'data' / 'modulos'
    mods_dir.mkdir(parents=True, exist_ok=True)

    index = []
    total_qs = 0
    for mod, questions in sorted(mods.items()):
        path = mods_dir / f'{mod}.json'
        path.write_text(
            json.dumps(questions, ensure_ascii=False, separators=(',',':')),
            encoding='utf-8'
        )
        size_kb = path.stat().st_size / 1024
        print(f"  {mod}: {len(questions)} preguntas — {size_kb:.1f} KB")
        index.append({'modulo': mod, 'total': len(questions)})
        total_qs += len(questions)

    # Write index.json
    idx_path = ROOT / 'data' / 'index.json'
    idx_path.write_text(
        json.dumps(index, ensure_ascii=False, separators=(',',':')),
        encoding='utf-8'
    )

    print(f"\n✓ {total_qs} preguntas en {len(index)} módulos")
    print(f"  index.json: {idx_path.stat().st_size / 1024:.1f} KB")

if __name__ == '__main__':
    convert()